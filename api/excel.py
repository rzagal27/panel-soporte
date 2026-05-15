from http.server import BaseHTTPRequestHandler
import json, io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

AZUL_OSC = "1F3864"
AZUL_MED = "2564CF"
AZUL_CLAR = "D6E4F7"
GRIS = "F2F2F2"
VERDE = "C6EFCE"
BLANCO = "FFFFFF"
AMARILLO = "FFF2CC"
EBF = "EBF3FB"

def fill(c): return PatternFill("solid", fgColor=c)

def fnt(bold=False, color="000000", size=11, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def bdr_all(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bdr_thick():
    s = Side(style="medium", color="1F3864")
    t = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def bdr_outer():
    s = Side(style="medium", color="1F3864")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_outer_border(ws, min_row, max_row, min_col, max_col):
    """Apply a medium border around a rectangular range."""
    from openpyxl.styles import Border, Side
    thick = Side(style="medium", color="1F3864")
    thin  = Side(style="thin",   color="BFBFBF")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            r, c = cell.row, cell.column
            left   = thick if c == min_col  else thin
            right  = thick if c == max_col  else thin
            top    = thick if r == min_row  else thin
            bottom = thick if r == max_row  else thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def generar_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Cotizacion_{data.get('citaServicio','')[:20]}"

    for col, w in [("A",6),("B",40),("C",9),("D",12),("E",16),("F",16)]:
        ws.column_dimensions[col].width = w

    # ── ROW 1: Título ────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"].value = "SOLICITUD DE COTIZACIÓN"
    ws["A1"].font  = Font(name="Calibri", bold=True, size=16, color=BLANCO)
    ws["A1"].fill  = fill(AZUL_OSC)
    ws["A1"].alignment = aln("center")
    ws.row_dimensions[1].height = 36

    # ── ROW 2: Barra azul ─────────────────────────────────────────────────────
    ws.merge_cells("A2:F2")
    ws["A2"].fill = fill(AZUL_MED)
    ws.row_dimensions[2].height = 6

    # ── ROWS 3-7: Datos proyecto ──────────────────────────────────────────────
    campos = [
        ("FM Group:",         data.get("fmGroup","")),
        ("Cita de Servicio:", data.get("citaServicio","")),
        ("Edificio:",         data.get("edificio","")),
        ("Dirección:",        data.get("direccion","")),
        ("Título:",           data.get("titulo","")),
    ]
    for i, (lbl, val) in enumerate(campos):
        r = 3 + i
        ws[f"A{r}"].value = lbl
        ws[f"A{r}"].font  = fnt(bold=True, color=AZUL_OSC)
        ws[f"A{r}"].fill  = fill(EBF)
        ws[f"A{r}"].alignment = aln("left")

        ws[f"B{r}"].fill = fill(EBF)

        ws.merge_cells(f"C{r}:F{r}")
        ws[f"C{r}"].value = val
        ws[f"C{r}"].font  = fnt()
        ws[f"C{r}"].fill  = fill(BLANCO)
        ws[f"C{r}"].alignment = aln("left")
        ws.row_dimensions[r].height = 18

    # Apply thin borders to project data section (A3:F7)
    apply_outer_border(ws, 3, 7, 1, 6)

    ws.row_dimensions[8].height = 8

    # ── ROW 9: Encabezado tabla ───────────────────────────────────────────────
    hdrs = ["N°", "Descripción / Partida", "Unidad", "Cantidad", "Precio Unit.", "Subtotal"]
    for col, h in zip("ABCDEF", hdrs):
        c = ws[f"{col}9"]
        c.value = h
        c.font  = Font(name="Calibri", bold=True, size=11, color=BLANCO)
        c.fill  = fill(AZUL_MED)
        c.alignment = aln("center")
        c.border = bdr_all("medium", "1F3864")
    ws.row_dimensions[9].height = 22

    # ── ROWS 10+: Partidas ────────────────────────────────────────────────────
    partidas = data.get("partidas", [])
    FIRST = 10
    LAST  = max(FIRST, FIRST + len(partidas) - 1)

    for i, p in enumerate(partidas):
        r  = FIRST + i
        bg = BLANCO if i % 2 == 0 else AZUL_CLAR
        ws.row_dimensions[r].height = 18

        ws[f"A{r}"].value = i + 1
        ws[f"A{r}"].font  = fnt(bold=True, color=AZUL_MED, size=10)
        ws[f"A{r}"].fill  = fill(bg)
        ws[f"A{r}"].alignment = aln("center")

        ws[f"B{r}"].value = p.get("descripcion","")
        ws[f"B{r}"].fill  = fill(bg)
        ws[f"B{r}"].alignment = aln("left", wrap=True)

        ws[f"C{r}"].value = p.get("unidad","un")
        ws[f"C{r}"].fill  = fill(bg)
        ws[f"C{r}"].alignment = aln("center")

        ws[f"D{r}"].value  = float(p.get("cantidad") or 0)
        ws[f"D{r}"].fill   = fill(AMARILLO)
        ws[f"D{r}"].alignment = aln("right")
        ws[f"D{r}"].number_format = "#,##0.00"

        ws[f"E{r}"].value  = float(p.get("precioUnitario") or 0)
        ws[f"E{r}"].fill   = fill(AMARILLO)
        ws[f"E{r}"].alignment = aln("right")
        ws[f"E{r}"].number_format = '"$"#,##0'

        ws[f"F{r}"].value  = f"=D{r}*E{r}"
        ws[f"F{r}"].fill   = fill(bg)
        ws[f"F{r}"].alignment = aln("right")
        ws[f"F{r}"].number_format = '"$"#,##0'
        ws[f"F{r}"].font   = fnt(bold=True)

    # Apply borders to items table
    apply_outer_border(ws, 9, LAST, 1, 6)

    # ── TOTALES ───────────────────────────────────────────────────────────────
    T        = LAST + 2
    gg_row   = T + 1
    uti_row  = T + 2
    neto_row = T + 3
    iva_row  = T + 4
    tot_row  = T + 5

    gg_pct  = float(data.get("gg",  10)) / 100
    uti_pct = float(data.get("uti", 10)) / 100

    def set_tot(row, lbl, formula, bg, is_total=False):
        ws.row_dimensions[row].height = 20
        for col in "ABCD":
            ws[f"{col}{row}"].fill = fill(BLANCO)

        e = ws[f"E{row}"]
        e.value = lbl
        e.font  = Font(name="Calibri", bold=True,
                       size=13 if is_total else 11,
                       color="107C10" if is_total else AZUL_OSC)
        e.fill  = fill(bg)
        e.alignment = aln("right")
        e.border = bdr_all()

        f = ws[f"F{row}"]
        f.value = formula
        f.font  = Font(name="Calibri", bold=True,
                       size=14 if is_total else 11,
                       color="107C10" if is_total else "000000")
        f.fill  = fill(bg)
        f.alignment = aln("right")
        f.border = bdr_all("medium","1F3864") if is_total else bdr_all()
        f.number_format = '"$"#,##0'

    # GG% editable
    ws[f"D{gg_row}"].value = gg_pct
    ws[f"D{gg_row}"].fill  = fill(AMARILLO)
    ws[f"D{gg_row}"].font  = fnt(bold=True, color="CC0000")
    ws[f"D{gg_row}"].alignment = aln("center")
    ws[f"D{gg_row}"].border = bdr_all()
    ws[f"D{gg_row}"].number_format = '0.0%'

    # UTI% editable
    ws[f"D{uti_row}"].value = uti_pct
    ws[f"D{uti_row}"].fill  = fill(AMARILLO)
    ws[f"D{uti_row}"].font  = fnt(bold=True, color="CC0000")
    ws[f"D{uti_row}"].alignment = aln("center")
    ws[f"D{uti_row}"].border = bdr_all()
    ws[f"D{uti_row}"].number_format = '0.0%'

    # IVA fijo
    ws[f"D{iva_row}"].value = 0.19
    ws[f"D{iva_row}"].fill  = fill(GRIS)
    ws[f"D{iva_row}"].font  = fnt(bold=True, color=AZUL_OSC)
    ws[f"D{iva_row}"].alignment = aln("center")
    ws[f"D{iva_row}"].border = bdr_all()
    ws[f"D{iva_row}"].number_format = '0%'

    set_tot(T,        "Subtotal Neto:", f"=SUM(F{FIRST}:F{LAST})",          GRIS)
    set_tot(gg_row,   "GG:",           f"=F{T}*D{gg_row}",                  GRIS)
    set_tot(uti_row,  "UTI:",          f"=F{T}*D{uti_row}",                 GRIS)
    set_tot(neto_row, "Neto:",         f"=F{T}+F{gg_row}+F{uti_row}",       GRIS)
    set_tot(iva_row,  "IVA:",          f"=F{neto_row}*D{iva_row}",          GRIS)
    set_tot(tot_row,  "TOTAL:",        f"=F{neto_row}+F{iva_row}",          VERDE, True)

    # Leyenda
    leg = tot_row + 2
    ws.merge_cells(f"A{leg}:F{leg}")
    ws[f"A{leg}"].value = "Las celdas en amarillo son editables (Cantidad, Precio Unitario, GG% y UTI%)"
    ws[f"A{leg}"].font  = Font(name="Calibri", size=9, italic=True, color="7F7F7F")
    ws[f"A{leg}"].alignment = aln("center")

    # ── SECCIÓN CONTRATISTA ───────────────────────────────────────────────────
    CS = leg + 2
    ws.merge_cells(f"A{CS}:F{CS}")
    ws[f"A{CS}"].value = "DATOS DEL CONTRATISTA"
    ws[f"A{CS}"].font  = Font(name="Calibri", bold=True, size=12, color=BLANCO)
    ws[f"A{CS}"].fill  = fill(AZUL_OSC)
    ws[f"A{CS}"].alignment = aln("center")
    ws.row_dimensions[CS].height = 22

    for i, lbl in enumerate(["Nombre Empresa:", "RUT Empresa:", "Fecha:"]):
        r = CS + 2 + i
        ws[f"A{r}"].value = lbl
        ws[f"A{r}"].font  = fnt(bold=True, color=AZUL_OSC)
        ws[f"A{r}"].fill  = fill(EBF)
        ws[f"A{r}"].alignment = aln("left")
        ws[f"A{r}"].border = bdr_all()

        ws.merge_cells(f"B{r}:F{r}")
        ws[f"B{r}"].fill  = fill(BLANCO)
        ws[f"B{r}"].border = bdr_all()
        ws.row_dimensions[r].height = 22

    # Apply outer border to contratista section
    apply_outer_border(ws, CS, CS + 4, 1, 6)

    # ── FIRMA (solo una) ──────────────────────────────────────────────────────
    firma_r = CS + 7
    ws.row_dimensions[firma_r].height = 60
    ws.merge_cells(f"A{firma_r}:F{firma_r}")
    ws[f"A{firma_r}"].value = "FIRMA EMPRESA"
    ws[f"A{firma_r}"].font  = fnt(bold=True, color=AZUL_OSC, size=10)
    ws[f"A{firma_r}"].fill  = fill(EBF)
    ws[f"A{firma_r}"].alignment = aln("center", "bottom")
    ws[f"A{firma_r}"].border = bdr_all("medium","1F3864")

    # Nota obligatorio
    nota_r = firma_r + 1
    ws.merge_cells(f"A{nota_r}:F{nota_r}")
    ws[f"A{nota_r}"].value = "* Firma y RUT empresa son OBLIGATORIOS para validar esta cotización"
    ws[f"A{nota_r}"].font  = Font(name="Calibri", size=9, italic=True, color="FF0000", bold=True)
    ws[f"A{nota_r}"].alignment = aln("center")

    # ── CONFIGURACIÓN HOJA ────────────────────────────────────────────────────
    # NO freeze panes
    ws.print_area = f"A1:F{nota_r}"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.orientation = "portrait"
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_POST(self):
        import base64
        try:
            length = int(self.headers.get("Content-Length", 0))
            body   = self.rfile.read(length)
            data   = json.loads(body)

            xlsx_bytes = generar_excel(data)
            b64 = base64.b64encode(xlsx_bytes).decode("utf-8")

            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps({"file": b64}).encode())
        except Exception as e:
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps({"error": str(e)}).encode())
